from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from controller.ExcelController import ExcelController
from model.ExcelModel import ExcelModel
from view.ReceiptView import ReceiptView

# work_book: Workbook = load_workbook(filename="ReceiptData.xlsx")
# work_sheet: Worksheet = work_book.active
# sheet_names: list = work_book.sheetnames
# print(sheet_names)
# property_row = work_sheet[2]
# needed_property = []
# student_data = []
#
# for property_in_row in property_row:
#     if property_in_row.value == "姓名" or \
#             property_in_row.value == "月費" or \
#             property_in_row.value == "保險(半年收一次)" or \
#             property_in_row.value == "其他(延拖費)" or \
#             property_in_row.value == "幼兒屬性" or \
#             property_in_row.value == "班別" or \
#             property_in_row.value == "請假減收":
#         needed_property.append(property_in_row.column - 1)
#
# print(needed_property)
#
# for student_property_column in needed_property:
#     student_data.append(work_sheet[convert_col_and_row_to_position(col_number=student_property_column , row_number=3)].value)
# student_data.append(work_book.sheetnames[0])
#
# print(Student(*student_data))
# above code is just note


controller = ExcelController()
model = ExcelModel()
view = ReceiptView(controller, model)

# need open the finished file function
