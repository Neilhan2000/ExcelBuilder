from openpyxl.reader.excel import load_workbook

from mapper.DataMapper import convert_col_and_row_to_position
from model.LoadTextModel import LoadTextModel
from model.LoadExcelModel import LoadExcelModel
import aiofiles
from model.Result import Result, Success, Error
from model.dataclass.Fee import Fee
from model.dataclass.Student import Student

note_path = "note.txt"


class ExcelModel(LoadTextModel, LoadExcelModel):
    _instance = None

    # implement singleton pattern
    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super().__new__(cls, *args, **kwargs)
        return cls._instance

    async def read_all_text_from_note(self) -> Result:
        try:
            async with aiofiles.open(file=note_path, mode="r", encoding="utf-8-sig") as file:
                cleaned_text_list = [item.rstrip('\n').split(':', 1)[1] for item in await file.readlines()]
                result = Fee(*cleaned_text_list)

                print(f"Read result successfully: data = \n{result}")
                return Success(data=result)

        except Exception as error:
            if file:
                await file.close()
            print(f"Exception: {error}")
            return Error(exception=error)

        finally:
            if file:
                await file.close()
                print("File was closed successfully")

    # TODO: load student from excel
    async def read_all_data_from_excel(self) -> Result:
        try:
            work_book = await load_workbook(filename="ReceiptData.xlsx")
            work_sheet = work_book.active
            property_row = work_sheet[2]
            needed_property = []
            student_data = []

            for property_in_row in property_row:
                if property_in_row.value == "姓名" or \
                        property_in_row.value == "月費" or \
                        property_in_row.value == "保險(半年收一次)" or \
                        property_in_row.value == "其他(延拖費)" or \
                        property_in_row.value == "幼兒屬性" or \
                        property_in_row.value == "班別" or \
                        property_in_row.value == "請假減收":
                    needed_property.append(property_in_row.column - 1)

            print(needed_property)

            for student_property_column in needed_property:
                student_data.append(work_sheet[convert_col_and_row_to_position(col_number=student_property_column , row_number=3)].value)
            student_data.append(work_book.sheetnames[0])

            result = Student(*student_data)

            print(f"Read result successfully: data = \n{result}")
            return Success(data=result)

        except Exception as error:
            print(f"Exception: {error}")
            return Error(exception=error)