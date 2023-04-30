from openpyxl.styles import Font, Alignment, Border
from openpyxl.worksheet.worksheet import Worksheet


def init_title_item(
        work_sheet: Worksheet,
        title: str,
        font_style: Font = Font(),
        alignment: Alignment = Alignment(),
        border: Border = Border(),
        border_row: int = 1,
        is_merged_title: bool = False,
        start_position: str = "A1",
        end_position: str = "A1"
):
    work_sheet.append([title])

    # initialize title cell
    if is_merged_title:
        work_sheet.merge_cells(f"${start_position}:${end_position}")

    work_sheet[start_position].font = font_style
    work_sheet[start_position].alignment = alignment

    # set merged cell border
    for i in range(convert_position_to_col_number(end_position) + 1):
        work_sheet[convert_col_and_row_to_position(col_number=i, row_number=border_row)].border = border


def init_multiple_merged_cell_item(
        work_sheet: Worksheet,
        position_value: list,
        split_position: list,
        font_style: Font = Font(),
        alignment: Alignment = Alignment(),
        specify_position_alignment=Alignment(),
        specify_position: str = None,
        border: Border = Border(),
):
    border_row = find_cell_digit(find_start_position(split_position[0]))

    # initialize all merged cells
    for cells_need_merged in split_position:
        start_position = find_start_position(cells_need_merged)

        work_sheet.merge_cells(cells_need_merged)
        work_sheet[start_position].value = position_value[split_position.index(cells_need_merged)]
        work_sheet[start_position].font = font_style
        work_sheet[start_position].alignment = alignment

    # find last element and use it to set up border
    for i in range(convert_position_to_col_number(find_start_position(split_position[0])), convert_position_to_col_number(find_end_position(split_position[-1])) + 1):
        work_sheet[convert_col_and_row_to_position(col_number=i, row_number=border_row)].border = border

    # set specify col
    if specify_position is not None:
        work_sheet[specify_position].alignment = specify_position_alignment


def init_three_column_fee_item(
        work_sheet: Worksheet,
        left_column_range: str,
        left_title: str,
        middle_columns_value: list,
        right_columns_value: list,
        left_column_font: Font = Font(),
        left_column_alignment=Alignment(),
        other_font: Font = Font(),
        middle_alignment=Alignment(),
        right_alignment=Alignment(),
        border=Border()
):
    start_position = find_start_position(left_column_range)
    start_column = start_position[0]
    start_row = find_cell_digit(start_position)
    total_row_number = int(find_cell_digit(find_end_position(left_column_range))) - int(
        find_cell_digit(start_position)) + 1
    middle_row_count = 0
    right_row_count = 0

    def get_left_position(increase_row: int) -> str:
        return convert_col_and_row_to_position(
            col_number=convert_column_str_to_int(column_str=start_column),
            row_number=start_row + increase_row
        )

    def get_middle_position(increase_row: int) -> str:
        return convert_col_and_row_to_position(
            col_number=convert_column_str_to_int(column_str=start_column) + 1,
            row_number=start_row + increase_row
        )

    def get_right_position(increase_row: int) -> str:
        return convert_col_and_row_to_position(
            col_number=convert_column_str_to_int(column_str=start_column) + 2,
            row_number=start_row + increase_row
        )

    # initialize left column
    work_sheet.merge_cells(left_column_range)
    work_sheet[start_position].value = left_title
    work_sheet[start_position].font = left_column_font
    work_sheet[start_position].alignment = left_column_alignment

    for i in range(total_row_number):
        left_position_cell = get_left_position(increase_row=i)
        work_sheet[left_position_cell].border = border

    # initialize middle column
    for middle_element in middle_columns_value:
        middle_position_cell = work_sheet[get_middle_position(increase_row=middle_row_count)]

        middle_position_cell.value = middle_element
        middle_position_cell.font = other_font
        middle_position_cell.alignment = middle_alignment
        middle_row_count += 1

    for i in range(total_row_number):
        middle_position_cell = get_middle_position(increase_row=i)
        work_sheet[middle_position_cell].border = border

    # initialize right column
    for middle_element in right_columns_value:
        right_position_cell = work_sheet[get_right_position(increase_row=right_row_count)]

        right_position_cell.value = middle_element
        right_position_cell.font = other_font
        right_position_cell.alignment = right_alignment
        right_row_count += 1

    for i in range(total_row_number):
        right_position_cell = get_right_position(increase_row=i)
        work_sheet[right_position_cell].border = border


def init_two_column_merged_row_item(
        work_sheet: Worksheet,
        left_split_row: list,
        right_split_row: list,
        left_column_values: list,
        right_column_values: list,
        left_alignment: Alignment = Alignment(),
        right_alignment: Alignment = Alignment(),
        text_font: Font = Font(),
        border: Border = Border
):
    left_row_count = 0
    right_row_count = 0

    # initialize left column
    for rows_need_merged in left_split_row:
        work_sheet.merge_cells(rows_need_merged)

        merged_start_cell = work_sheet[find_start_position(rows_need_merged)]
        merged_start_cell.alignment = left_alignment
        merged_start_cell.font = text_font
        merged_start_cell.value = left_column_values[left_row_count]
        left_row_count += 1

        merged_end_cell = work_sheet[find_end_position(rows_need_merged)]
        merged_start_cell.border = border
        merged_end_cell.border = border

    # initialize right column
    for rows_need_merged in right_split_row:
        work_sheet.merge_cells(rows_need_merged)

        merged_start_cell = work_sheet[find_start_position(rows_need_merged)]
        merged_start_cell.alignment = right_alignment
        merged_start_cell.font = text_font
        merged_start_cell.value = right_column_values[right_row_count]
        right_row_count += 1

        merged_end_cell = work_sheet[find_end_position(rows_need_merged)]
        merged_start_cell.border = border
        merged_end_cell.border = border


def init_row_and_column_merged_item(
        work_sheet: Worksheet,
        merged_range: str,
        value: str,
        text_font: Font = Font(),
        text_alignment: Alignment = Alignment(),
        border: Border = Border()
):
    work_sheet.merge_cells(merged_range)

    merged_cell = work_sheet[find_start_position(merged_range)]
    merged_cell.value = value
    merged_cell.font = text_font
    merged_cell.alignment = text_alignment

    def get_position(col_num: int, row_num: int) -> str:
        return convert_col_and_row_to_position(
            col_number=col_num,
            row_number=row_num
        )

    # recursive function to traverse all rows
    def traverse_all_rows_and_set_border(col_list: list, row_list: list, col_num, row_num):
        if row_num <= row_list[-1]:
            work_sheet[
                get_position(
                    col_num=col_num,
                    row_num=row_num
                )
            ].border = border

            traverse_all_rows_and_set_border(col_list=col_list, row_list=row_list, col_num=col_num, row_num=row_num + 1)

    start_column_number = convert_position_to_col_number(position=find_start_position(merged_range))
    end_column_number = convert_position_to_col_number(position=find_end_position(merged_range))
    start_row = find_cell_digit(find_start_position(merged_range))
    end_row = find_cell_digit(find_end_position(merged_range))

    # traverse all cell and set border
    column_list = []
    for column_number in range(start_column_number, end_column_number + 1):
        column_list.append(column_number)

    row_list = []
    for row_number in range(start_row, end_row + 1):
        row_list.append(row_number)

    for column in column_list:
        traverse_all_rows_and_set_border(column_list, row_list, column, row_list[0])


def convert_col_and_row_to_position(col_number: int, row_number: int) -> str:  # make it private
    match col_number:
        case 0:
            return f"A{row_number}"
        case 1:
            return f"B{row_number}"
        case 2:
            return f"C{row_number}"
        case 3:
            return f"D{row_number}"
        case 4:
            return f"E{row_number}"
        case 5:
            return f"F{row_number}"
        case 6:
            return f"G{row_number}"
        case 7:
            return f"H{row_number}"
        case 8:
            return f"I{row_number}"
        case 9:
            return f"J{row_number}"


def convert_position_to_col_number(position: str) -> int:
    if position.__contains__("A"):
        return 0
    if position.__contains__("B"):
        return 1
    if position.__contains__("C"):
        return 2
    if position.__contains__("D"):
        return 3
    if position.__contains__("E"):
        return 4
    if position.__contains__("F"):
        return 5
    if position.__contains__("G"):
        return 6
    if position.__contains__("H"):
        return 7
    if position.__contains__("I"):
        return 8
    if position.__contains__("J"):
        return 9


def convert_column_str_to_int(column_str) -> int:
    match column_str:
        case "A":
            return 0
        case "B":
            return 1
        case "C":
            return 2
        case "D":
            return 3
        case "E":
            return 4
        case "F":
            return 5
        case "G":
            return 6
        case "H":
            return 7
        case "I":
            return 8
        case "J":
            return 9


def find_start_position(cells_str: str):
    return cells_str.split(":")[0]


def find_end_position(cells_str: str):
    return cells_str.split(":")[-1]


def find_cell_digit(cell: str) -> int:
    return int("".join(filter(str.isdigit, cell)))
